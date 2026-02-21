#!/usr/bin/env python3
"""
Build the unified ingredients reference database from all bulk data sources.

Sources:
1. Seed ingredient list (from FOOD_DB extraction)
2. FDA bulk adverse events (148k events → aggregate by ingredient name)
3. FDA bulk enforcement/recalls (28k recalls → aggregate by ingredient name)
4. EFSA OpenFoodTox (substances + reference values → ADI, NOAEL, hazards)
5. IARC Monographs (carcinogen classifications from PDF)

Output: scripts/d1-ingredients-ref-unified.sql

Usage:
    python3 scripts/build-reference-db.py

Then import:
    npx wrangler d1 execute consumer-truth-ingredients-ref --remote --file=scripts/d1-ingredients-ref-unified.sql
"""

import json
import os
import re
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

try:
    import pdfplumber
except ImportError:
    print("Installing pdfplumber...")
    os.system("pip install pdfplumber")
    import pdfplumber

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BULK_DIR = os.path.join(BASE_DIR, "bulk-data")
SEED_FILE = os.path.join(BASE_DIR, "ingredient-seed-list.json")
OUTPUT_FILE = os.path.join(BASE_DIR, "d1-ingredients-ref-unified.sql")
SQL_BATCH_SIZE = 20


def escape_sql(s: str) -> str:
    if not s:
        return ""
    return s.replace("'", "''")


def decode_xlsx_str(s) -> str:
    """Decode XML-encoded strings from EFSA xlsx files."""
    if not s or not isinstance(s, str):
        return str(s) if s else ""
    # Replace XML hex escapes: _x0028_ → (
    def hex_replace(m):
        code = int(m.group(1), 16)
        return chr(code)
    s = re.sub(r'_x([0-9A-Fa-f]{4})_', hex_replace, s)
    return s.strip()


# ─── 1. LOAD SEED LIST ───────────────────────────────────────────────────────

def load_seed_list() -> list[str]:
    print("Loading seed ingredient list...")
    with open(SEED_FILE) as f:
        names = json.load(f)
    print(f"  {len(names)} ingredient names")
    return names


# ─── 2. PARSE FDA ADVERSE EVENTS ─────────────────────────────────────────────

def parse_fda_events() -> dict[str, int]:
    """Parse FDA bulk adverse events, count mentions per ingredient-like term."""
    events_file = os.path.join(BULK_DIR, "fda-events", "food-event-0001-of-0001.json")
    if not os.path.exists(events_file):
        print("  FDA events file not found, skipping")
        return {}

    print("Parsing FDA adverse events (104MB)...")
    with open(events_file) as f:
        data = json.load(f)

    results = data.get("results", [])
    print(f"  {len(results)} total events")

    # Count by industry_name (product category) and reactions
    name_counts: dict[str, int] = defaultdict(int)

    for event in results:
        products = event.get("products", [])
        for prod in products:
            name = (prod.get("industry_name") or "").strip().lower()
            if name and len(name) >= 3:
                name_counts[name] += 1

    print(f"  {len(name_counts)} unique product/industry names with events")
    return dict(name_counts)


# ─── 3. PARSE FDA ENFORCEMENT/RECALLS ────────────────────────────────────────

def parse_fda_recalls() -> dict[str, dict]:
    """Parse FDA bulk enforcement data, extract recall counts and recent reasons."""
    recalls_file = os.path.join(BULK_DIR, "fda-enforcement", "food-enforcement-0001-of-0001.json")
    if not os.path.exists(recalls_file):
        print("  FDA enforcement file not found, skipping")
        return {}

    print("Parsing FDA enforcement/recalls (38MB)...")
    with open(recalls_file) as f:
        data = json.load(f)

    results = data.get("results", [])
    print(f"  {len(results)} total recalls")

    # Index by keywords in reason_for_recall
    # We'll do fuzzy matching against our seed list later
    recall_reasons: list[dict] = []
    for r in results:
        reason = r.get("reason_for_recall", "")
        classification = r.get("classification", "Unknown")
        status = r.get("status", "Unknown")
        recall_reasons.append({
            "reason": reason[:200],
            "classification": classification,
            "status": status,
            "reason_lower": reason.lower(),
        })

    print(f"  Indexed {len(recall_reasons)} recalls for matching")
    return {"recalls": recall_reasons, "total": len(recall_reasons)}


def match_fda_recalls(seed_names: list[str], recall_data: dict) -> dict[str, dict]:
    """Match ingredient names against FDA recall reasons."""
    if not recall_data:
        return {}

    recalls = recall_data.get("recalls", [])
    if not recalls:
        return {}

    print("Matching ingredient names against FDA recalls...")
    results: dict[str, dict] = {}

    # Only match ingredients that are chemical/additive-like (not generic food words)
    # Pre-filter to ingredients with 2+ words or known chemical patterns
    for name in seed_names:
        count = 0
        recent = []
        for r in recalls:
            if name in r["reason_lower"]:
                count += 1
                if len(recent) < 3:
                    recent.append({
                        "reason": r["reason"],
                        "classification": r["classification"],
                        "status": r["status"],
                    })

        if count > 0:
            results[name] = {
                "total": count,
                "recent": recent,
            }

    print(f"  {len(results)} ingredients matched to recalls")
    return results


# ─── 4. PARSE EFSA OPENFOODTOX ───────────────────────────────────────────────

def parse_efsa_substances() -> dict[str, dict]:
    """Parse EFSA substance characterisation for CAS numbers and molecular info."""
    xlsx_file = os.path.join(BULK_DIR, "efsa-substances.xlsx")
    if not os.path.exists(xlsx_file):
        print("  EFSA substances file not found, skipping")
        return {}

    print("Parsing EFSA substances...")
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    ws = wb.active

    results: dict[str, dict] = {}
    header = None

    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [decode_xlsx_str(c) for c in row]
            continue

        vals = [decode_xlsx_str(c) for c in row]
        if len(vals) < 4:
            continue

        substance = vals[0].lower().strip()
        cas = vals[3].strip() if vals[3] else ""
        formula = vals[5].strip() if len(vals) > 5 and vals[5] else ""

        if substance and len(substance) >= 3:
            results[substance] = {
                "cas_number": cas if cas and re.match(r'^\d{1,7}-\d{2}-\d$', cas) else "",
                "molecular_formula": formula,
            }

    wb.close()
    print(f"  {len(results)} EFSA substances loaded")
    return results


def parse_efsa_reference_values() -> dict[str, dict]:
    """Parse EFSA reference values for ADI and NOAEL."""
    xlsx_file = os.path.join(BULK_DIR, "efsa-reference-values.xlsx")
    if not os.path.exists(xlsx_file):
        print("  EFSA reference values file not found, skipping")
        return {}

    print("Parsing EFSA reference values (ADI/NOAEL)...")
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    ws = wb.active

    results: dict[str, dict] = {}
    header = None

    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [decode_xlsx_str(c) for c in row]
            continue

        vals = [decode_xlsx_str(c) for c in row]
        if len(vals) < 8:
            continue

        substance = vals[0].lower().strip()
        assessment = vals[4].strip() if vals[4] else ""
        value = vals[6]
        unit = vals[7].strip() if vals[7] else ""
        year = vals[2]

        if not substance or len(substance) < 3:
            continue

        if substance not in results:
            results[substance] = {}

        assessment_lower = assessment.lower()
        if "adi" in assessment_lower:
            results[substance]["adi"] = f"{value} {unit}" if value else ""
            results[substance]["year"] = year
        elif "noael" in assessment_lower or "noel" in assessment_lower:
            results[substance]["noael"] = f"{value} {unit}" if value else ""
            results[substance]["year"] = year
        elif "bmdl" in assessment_lower or "benchmark" in assessment_lower:
            results[substance]["noael"] = f"BMDL: {value} {unit}" if value else ""
            results[substance]["year"] = year

    wb.close()

    # Filter to only entries with actual values
    results = {k: v for k, v in results.items() if v.get("adi") or v.get("noael")}
    print(f"  {len(results)} substances with ADI/NOAEL values")
    return results


# ─── 5. PARSE IARC CARCINOGEN CLASSIFICATIONS ────────────────────────────────

def parse_iarc_pdf() -> dict[str, dict]:
    """Load pre-parsed IARC classifications (parsed from PDF via text extraction)."""
    parsed_file = os.path.join(BULK_DIR, "iarc-parsed.json")

    # If pre-parsed file exists, use it directly
    if os.path.exists(parsed_file):
        print("Loading pre-parsed IARC classifications...")
        with open(parsed_file) as f:
            results = json.load(f)
        print(f"  {len(results)} IARC classifications loaded")
        groups = defaultdict(int)
        for v in results.values():
            groups[v["group"]] += 1
        for g in sorted(groups.keys()):
            print(f"    Group {g}: {groups[g]} agents")
        return results

    # Fallback: parse PDF with text extraction
    pdf_file = os.path.join(BULK_DIR, "iarc-classifications.pdf")
    if not os.path.exists(pdf_file):
        print("  IARC data not found, skipping")
        return {}

    print("Parsing IARC carcinogen classifications PDF...")
    results: dict[str, dict] = {}

    line_pattern = re.compile(
        r'^(\d{3,7}-\d{2}-\d\*?)?\s*'
        r'(.+?)\s+'
        r'(1|2A|2B|3)\s+'
        r'(\d[\d\w\s,]*?)\s+'
        r'(\d{4})\s*$'
    )

    group_desc = {
        "1": "Carcinogenic to humans",
        "2A": "Probably carcinogenic to humans",
        "2B": "Possibly carcinogenic to humans",
        "3": "Not classifiable as to carcinogenicity",
    }

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                line = line.strip()
                if not line or line.startswith('CAS No.') or line.startswith('Agents Classified'):
                    continue
                if line.startswith('(NB:') or line.startswith('supporting') or re.match(r'^\d{1,2}$', line):
                    continue

                m = line_pattern.match(line)
                if m:
                    cas = (m.group(1) or "").strip().rstrip('*')
                    agent = m.group(2).strip()
                    group = m.group(3)
                    agent_clean = re.sub(r'\s*\(NB:.*', '', agent).strip()
                    agent_lower = agent_clean.lower()

                    if len(agent_lower) >= 3:
                        results[agent_lower] = {
                            "group": group,
                            "description": group_desc.get(group, ""),
                            "agent_name": agent_clean,
                            "cas": cas if cas and re.match(r'^\d{3,7}-\d{2}-\d$', cas) else "",
                        }

    # Save for next time
    with open(parsed_file, 'w') as f:
        json.dump(results, f)

    print(f"  {len(results)} IARC classifications loaded")
    groups = defaultdict(int)
    for v in results.values():
        groups[v["group"]] += 1
    for g in sorted(groups.keys()):
        print(f"    Group {g}: {groups[g]} agents")
    return results


# ─── 6. MERGE AND GENERATE SQL ───────────────────────────────────────────────

def build_unified_sql(
    seed_names: list[str],
    fda_events: dict[str, int],
    fda_recalls_matched: dict[str, dict],
    efsa_substances: dict[str, dict],
    efsa_ref_values: dict[str, dict],
    iarc_data: dict[str, dict],
):
    """Merge all data sources and generate unified SQL file."""
    print(f"\nBuilding unified SQL for {len(seed_names)} seed ingredients...")

    # Build a lookup for all seed names
    all_names = set(n.lower() for n in seed_names)

    # Also add all IARC agents and EFSA substances that aren't in seed list
    for name in iarc_data:
        all_names.add(name)
    for name in efsa_ref_values:
        all_names.add(name)

    print(f"  Total unique names (seed + IARC + EFSA): {len(all_names)}")

    # Build merged records
    records: list[dict] = []

    for name in sorted(all_names):
        rec = {
            "name": name,
            "name_original": name,
            "cas_number": None,
            "molecular_formula": None,
            "fda_adverse_event_count": 0,
            "fda_recall_count": 0,
            "fda_recent_recalls": "[]",
            "efsa_adi": None,
            "efsa_noael": None,
            "efsa_hazard": None,
            "efsa_evaluation_year": None,
            "iarc_group": None,
            "iarc_description": None,
            "iarc_agent_name": None,
            "safety_concerns": [],
        }

        # FDA events
        event_count = fda_events.get(name, 0)
        rec["fda_adverse_event_count"] = event_count
        if event_count > 100:
            rec["safety_concerns"].append(f"High FDA adverse event count ({event_count})")

        # FDA recalls
        recalls = fda_recalls_matched.get(name)
        if recalls:
            rec["fda_recall_count"] = recalls["total"]
            rec["fda_recent_recalls"] = json.dumps(recalls["recent"])
            if recalls["total"] > 0:
                rec["safety_concerns"].append(f"FDA recalls: {recalls['total']}")

        # EFSA substances (CAS, formula)
        efsa_sub = efsa_substances.get(name, {})
        if efsa_sub.get("cas_number"):
            rec["cas_number"] = efsa_sub["cas_number"]
        if efsa_sub.get("molecular_formula"):
            rec["molecular_formula"] = efsa_sub["molecular_formula"]

        # EFSA reference values (ADI, NOAEL)
        efsa_ref = efsa_ref_values.get(name, {})
        if efsa_ref.get("adi"):
            rec["efsa_adi"] = efsa_ref["adi"]
        if efsa_ref.get("noael"):
            rec["efsa_noael"] = efsa_ref["noael"]
        if efsa_ref.get("year"):
            rec["efsa_evaluation_year"] = efsa_ref["year"]

        # IARC
        iarc = iarc_data.get(name)
        if iarc:
            rec["iarc_group"] = iarc["group"]
            rec["iarc_description"] = iarc["description"]
            rec["iarc_agent_name"] = iarc["agent_name"]
            if iarc.get("cas") and not rec["cas_number"]:
                rec["cas_number"] = iarc["cas"]

            if iarc["group"] in ("1", "2A"):
                rec["safety_concerns"].append(f"IARC Group {iarc['group']}: {iarc['description']}")

        records.append(rec)

    # Filter: only keep records that have at least SOME data beyond just a name
    enriched = [r for r in records if (
        r["cas_number"] or
        r["fda_adverse_event_count"] > 0 or
        r["fda_recall_count"] > 0 or
        r["efsa_adi"] or
        r["efsa_noael"] or
        r["iarc_group"] or
        r["molecular_formula"]
    )]

    # Also keep all seed list items (they'll get PubChem data later via API)
    seed_set = set(n.lower() for n in seed_names)
    for r in records:
        if r["name"] in seed_set and r not in enriched:
            enriched.append(r)

    enriched.sort(key=lambda r: r["name"])
    print(f"  {len(enriched)} records with data (out of {len(records)} total)")

    # Generate SQL
    with open(OUTPUT_FILE, "w") as out:
        out.write("-- Unified ingredients reference database\n")
        out.write("-- Auto-generated by build-reference-db.py\n")
        out.write(f"-- Sources: FDA bulk ({len(fda_events)} event categories, {fda_recalls_matched.get('__total', 0)} recall matches), ")
        out.write(f"EFSA ({len(efsa_substances)} substances, {len(efsa_ref_values)} ref values), ")
        out.write(f"IARC ({len(iarc_data)} classifications)\n\n")

        batch = []
        for rec in enriched:
            cas = f"'{escape_sql(rec['cas_number'])}'" if rec["cas_number"] else "NULL"
            mf = f"'{escape_sql(rec['molecular_formula'])}'" if rec["molecular_formula"] else "NULL"
            efsa_adi = f"'{escape_sql(rec['efsa_adi'])}'" if rec["efsa_adi"] else "NULL"
            efsa_noael = f"'{escape_sql(rec['efsa_noael'])}'" if rec["efsa_noael"] else "NULL"
            efsa_hazard = f"'{escape_sql(rec['efsa_hazard'])}'" if rec["efsa_hazard"] else "NULL"
            efsa_year = str(int(rec["efsa_evaluation_year"])) if rec["efsa_evaluation_year"] else "NULL"
            iarc_group = f"'{escape_sql(rec['iarc_group'])}'" if rec["iarc_group"] else "NULL"
            iarc_desc = f"'{escape_sql(rec['iarc_description'])}'" if rec["iarc_description"] else "NULL"
            iarc_agent = f"'{escape_sql(rec['iarc_agent_name'])}'" if rec["iarc_agent_name"] else "NULL"
            concerns = f"'{escape_sql(json.dumps(rec['safety_concerns']))}'" if rec["safety_concerns"] else "'[]'"

            values = (
                f"('{escape_sql(rec['name'])}', '{escape_sql(rec['name_original'])}', "
                f"{cas}, NULL, {mf}, NULL, NULL, "
                f"{rec['fda_adverse_event_count']}, {rec['fda_recall_count']}, "
                f"'{escape_sql(rec['fda_recent_recalls'])}', datetime('now'), "
                f"{efsa_adi}, {efsa_noael}, {efsa_hazard}, {efsa_year}, "
                f"{iarc_group}, {iarc_desc}, {iarc_agent}, "
                f"NULL, 0, NULL, '[]', NULL, "
                f"0, '[]', {concerns})"
            )
            batch.append(values)

            if len(batch) >= SQL_BATCH_SIZE:
                _flush_batch(out, batch)
                batch = []

        if batch:
            _flush_batch(out, batch)

    file_size = os.path.getsize(OUTPUT_FILE) / 1024 / 1024
    print(f"\n  Output: {OUTPUT_FILE} ({file_size:.1f} MB)")
    print(f"  Total records: {len(enriched)}")

    # Stats summary
    stats = {
        "with_cas": sum(1 for r in enriched if r["cas_number"]),
        "with_fda_events": sum(1 for r in enriched if r["fda_adverse_event_count"] > 0),
        "with_fda_recalls": sum(1 for r in enriched if r["fda_recall_count"] > 0),
        "with_efsa_adi": sum(1 for r in enriched if r["efsa_adi"]),
        "with_iarc": sum(1 for r in enriched if r["iarc_group"]),
        "with_concerns": sum(1 for r in enriched if r["safety_concerns"]),
    }
    print(f"\n  Enrichment stats:")
    for k, v in stats.items():
        print(f"    {k}: {v}")


def _flush_batch(out, batch):
    cols = (
        "name, name_original, cas_number, pubchem_cid, molecular_formula, "
        "molecular_weight, iupac_name, fda_adverse_event_count, fda_recall_count, "
        "fda_recent_recalls, last_fda_sync_at, efsa_adi, efsa_noael, efsa_hazard, "
        "efsa_evaluation_year, iarc_group, iarc_description, iarc_agent_name, "
        "e_number, eu_approved, eu_max_level, eu_food_categories, eu_restrictions, "
        "is_banned_anywhere, banned_in, safety_concerns"
    )
    out.write(f"INSERT OR IGNORE INTO ingredient_reference ({cols})\nVALUES\n  ")
    out.write(",\n  ".join(batch))
    out.write(";\n\n")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Building unified ingredients reference database")
    print("=" * 60)

    # Load all sources
    seed_names = load_seed_list()
    fda_events = parse_fda_events()
    fda_recall_data = parse_fda_recalls()
    fda_recalls_matched = match_fda_recalls(seed_names, fda_recall_data)
    efsa_substances = parse_efsa_substances()
    efsa_ref_values = parse_efsa_reference_values()
    iarc_data = parse_iarc_pdf()

    # Build unified SQL
    build_unified_sql(
        seed_names,
        fda_events,
        fda_recalls_matched,
        efsa_substances,
        efsa_ref_values,
        iarc_data,
    )

    print("\n" + "=" * 60)
    print("Done! Next steps:")
    print("  1. npx wrangler d1 execute consumer-truth-ingredients-ref --remote --file=scripts/d1-ingredients-ref-schema.sql")
    print("  2. npx wrangler d1 execute consumer-truth-ingredients-ref --remote --file=scripts/d1-ingredients-ref-unified.sql")
    print("  3. npx tsx scripts/import-pubchem.ts  (for PubChem CID/formula/weight/IUPAC)")
    print("  4. npx wrangler d1 execute consumer-truth-ingredients-ref --remote --file=scripts/d1-ingredients-ref-pubchem.sql")
    print("=" * 60)


if __name__ == "__main__":
    main()
