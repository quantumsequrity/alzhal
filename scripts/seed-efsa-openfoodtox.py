#!/usr/bin/env python3
"""
EFSA OpenFoodTox → Canonical Ingredient Graph seeder.

OpenFoodTox publishes the outcomes of EFSA's chemical risk assessments:
- Component substances (name + CAS)
- Studies linking substances to hazards (mutagenic, genotoxic, carcinogenic)
- Chemical assessments (ADI, NOAEL, ARfD) with risk values and units

Input:
    scripts/bulk-data/efsa-openfoodtox.xlsx
    Download: https://zenodo.org/records/8120114

Output:
    scripts/d1-regulatory-efsa.sql

Join path:
    COMPONENT (SUB_COM_ID, SUB_NAME, SUB_CASNUMBER)
      → STUDY (SUB_COM_ID → HAZARD_ID, IS_MUTAGENIC, IS_GENOTOXIC, IS_CARCINOGENIC)
      → CHEM_ASSESS (HAZARD_ID → ASSESSMENTTYPE, RISKVALUE, RISKUNIT)
"""

import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print('openpyxl required. Install with: pip install openpyxl', file=sys.stderr)
    sys.exit(2)

BASE = Path(__file__).resolve().parent
INPUT = BASE / 'bulk-data' / 'efsa-openfoodtox.xlsx'
OUTPUT = BASE / 'd1-regulatory-efsa.sql'

INGESTER = 'seed-efsa-openfoodtox'
VERSION = '1.0.0'
SOURCE_NAME = 'EFSA OpenFoodTox (Zenodo record 8120114, 2023)'
SOURCE_URL = 'https://zenodo.org/records/8120114'


def esc(s):
    if s is None or s == '':
        return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"


def num(v):
    if v is None or v == '':
        return 'NULL'
    try:
        f = float(v)
        if f != f: return 'NULL'
        return f'{f:g}'
    except (ValueError, TypeError):
        return 'NULL'


def slug(s):
    s = (s or '').lower()
    s = re.sub(r'\([^)]*\)', '', s)
    s = re.sub(r'[^\w\s-]', '', s)
    s = re.sub(r'[\s_-]+', '_', s).strip('_')
    return s[:80]


def normalize_alias(s):
    if not s: return ''
    d = unicodedata.normalize('NFD', s)
    stripped = ''.join(c for c in d if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', stripped).strip().lower()


CAS_RE = re.compile(r'^\d{2,7}-\d{2}-\d$')


def is_valid_cas(s):
    return bool(s and isinstance(s, str) and CAS_RE.match(s.strip()))


def main():
    if not INPUT.exists():
        print(f'Missing {INPUT}. Download: curl -L -o {INPUT.relative_to(Path.cwd())} '
              f'"https://zenodo.org/api/records/8120114/files/OpenFoodToxTX22809_2023.xlsx/content"',
              file=sys.stderr)
        sys.exit(2)

    print(f'Loading {INPUT.name}...')
    wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)

    # 1) COMPONENT: SUB_COM_ID → {name, cas}
    print('  reading COMPONENT...')
    components = {}
    ws = wb['COMPONENT']
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = {v: i for i, v in enumerate(row)}
            continue
        sub_com = row[headers['SUB_COM_ID']]
        name = row[headers['SUB_NAME']]
        cas = row[headers.get('SUB_CASNUMBER', -1)] if 'SUB_CASNUMBER' in headers else None
        if sub_com and name:
            components[sub_com] = {'name': str(name).strip(), 'cas': str(cas).strip() if cas else None}
    print(f'    {len(components)} components')

    # 2) STUDY: SUB_COM_ID → HAZARD_IDs + hazard flags
    print('  reading STUDY (this takes a moment)...')
    sub_to_hazard = defaultdict(set)
    sub_hazards = defaultdict(lambda: {'mutagenic': False, 'genotoxic': False, 'carcinogenic': False})
    ws = wb['STUDY']
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = {v: i for i, v in enumerate(row)}
            continue
        sub_com = row[headers['SUB_COM_ID']]
        hazard = row[headers.get('HAZARD_ID', -1)] if 'HAZARD_ID' in headers else None
        mut = row[headers.get('IS_MUTAGENIC', -1)] if 'IS_MUTAGENIC' in headers else None
        gen = row[headers.get('IS_GENOTOXIC', -1)] if 'IS_GENOTOXIC' in headers else None
        carc = row[headers.get('IS_CARCINOGENIC', -1)] if 'IS_CARCINOGENIC' in headers else None
        if sub_com and hazard:
            sub_to_hazard[sub_com].add(hazard)
        if sub_com:
            if mut and str(mut).strip().lower() in ('y', 'yes', 'true', '1'):
                sub_hazards[sub_com]['mutagenic'] = True
            if gen and str(gen).strip().lower() in ('y', 'yes', 'true', '1'):
                sub_hazards[sub_com]['genotoxic'] = True
            if carc and str(carc).strip().lower() in ('y', 'yes', 'true', '1'):
                sub_hazards[sub_com]['carcinogenic'] = True
    print(f'    {len(sub_to_hazard)} substances with hazard links, {len(sub_hazards)} with flags')

    # 3) CHEM_ASSESS: HAZARD_ID → list of {type, value, unit}
    print('  reading CHEM_ASSESS...')
    hazard_to_assessments = defaultdict(list)
    ws = wb['CHEM_ASSESS']
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = {v: i for i, v in enumerate(row)}
            continue
        hazard = row[headers['HAZARD_ID']]
        atype = row[headers.get('ASSESSMENTTYPE', -1)]
        qual = row[headers.get('RISKQUALIFIER', -1)]
        val = row[headers.get('RISKVALUE', -1)]
        unit = row[headers.get('RISKUNIT', -1)]
        population = row[headers.get('POPULATIONTEXT', -1)] if 'POPULATIONTEXT' in headers else None
        if hazard and atype and val is not None:
            hazard_to_assessments[hazard].append({
                'type': str(atype).strip(),
                'qual': str(qual).strip() if qual else '',
                'value': val,
                'unit': str(unit).strip() if unit else '',
                'population': str(population).strip() if population else '',
            })
    print(f'    {len(hazard_to_assessments)} hazards with assessments')

    # Build per-substance output
    snapshot = datetime.now(timezone.utc).date().isoformat()
    lines = []
    lines.append('-- ============================================================')
    lines.append(f'-- EFSA OpenFoodTox → CIG regulatory facts')
    lines.append(f'-- Generated: {datetime.now(timezone.utc).isoformat()}')
    lines.append(f'-- Ingester:  {INGESTER} v{VERSION}')
    lines.append(f'-- Source:    {SOURCE_URL}')
    lines.append('-- ============================================================')
    lines.append('')
    lines.append(f"INSERT INTO ingestion_run (ingester_name, ingester_version, status) "
                 f"VALUES ({esc(INGESTER)}, {esc(VERSION)}, 'running');")
    lines.append('')
    lines.append(
        f"INSERT INTO fact_evidence (source_name, source_url, source_section, snapshot_date, language, retrieved_by) "
        f"VALUES ({esc(SOURCE_NAME)}, {esc(SOURCE_URL)}, 'OpenFoodTox 2023 xlsx dump', {esc(snapshot)}, 'en', {esc(INGESTER + ':' + VERSION)}) "
        f"ON CONFLICT(source_url, snapshot_date) DO NOTHING;"
    )
    lines.append('')

    seen_canonical = set()
    ingredient_count = 0
    alias_count = 0
    fact_count = 0
    skipped = 0

    for sub_com, info in components.items():
        name = info['name']
        cas = info['cas']
        if not name or len(name) > 200:
            skipped += 1
            continue

        hazards = sub_to_hazard.get(sub_com, set())
        flags = sub_hazards.get(sub_com, {'mutagenic': False, 'genotoxic': False, 'carcinogenic': False})

        # Collect assessments
        assessments = []
        for h in hazards:
            assessments.extend(hazard_to_assessments.get(h, []))

        # Skip substances with no facts at all
        if not assessments and not any(flags.values()):
            skipped += 1
            continue

        # canonical_id
        if is_valid_cas(cas):
            canonical_id = f'CAS_{cas.replace("-", "_")}'
        else:
            canonical_id = f'NAME_{slug(name)}'
            if canonical_id == 'NAME_':
                continue

        # Upsert ingredient
        if canonical_id not in seen_canonical:
            seen_canonical.add(canonical_id)
            lines.append(
                f"INSERT INTO ingredient (canonical_id, primary_name, ingredient_class, cas_number, category, is_natural) "
                f"VALUES ({esc(canonical_id)}, {esc(name)}, 'substance', {esc(cas) if is_valid_cas(cas) else 'NULL'}, 'multi', 0) "
                f"ON CONFLICT(canonical_id) DO NOTHING;"
            )
            ingredient_count += 1

            lines.append(
                f"INSERT INTO ingredient_alias (canonical_id, alias, alias_normalized, alias_type, language_code, confidence, source) "
                f"VALUES ({esc(canonical_id)}, {esc(name)}, {esc(normalize_alias(name))}, 'synonym', 'en', 1.0, 'EFSA-OpenFoodTox') "
                f"ON CONFLICT(alias_normalized, alias_type, COALESCE(language_code, '')) DO NOTHING;"
            )
            alias_count += 1
            if is_valid_cas(cas):
                lines.append(
                    f"INSERT INTO ingredient_alias (canonical_id, alias, alias_normalized, alias_type, language_code, confidence, source) "
                    f"VALUES ({esc(canonical_id)}, {esc(cas)}, {esc(normalize_alias(cas))}, 'cas', NULL, 1.0, 'EFSA-OpenFoodTox') "
                    f"ON CONFLICT(alias_normalized, alias_type, COALESCE(language_code, '')) DO NOTHING;"
                )
                alias_count += 1

        # Emit facts
        # Hazard flags → classification facts
        hazard_flags = []
        if flags['carcinogenic']: hazard_flags.append('carcinogenic')
        if flags['genotoxic']:    hazard_flags.append('genotoxic')
        if flags['mutagenic']:    hazard_flags.append('mutagenic')
        if hazard_flags:
            status = f"EFSA hazard flags: {', '.join(hazard_flags)}"
            lines.append(
                f"INSERT OR IGNORE INTO regulatory_fact (canonical_id, jurisdiction, fact_type, status, regulation_ref, evidence_id) "
                f"VALUES ({esc(canonical_id)}, 'EFSA', 'classification', {esc(status)}, "
                f"'EFSA OpenFoodTox hazard flags', "
                f"(SELECT id FROM fact_evidence WHERE source_url = {esc(SOURCE_URL)} AND snapshot_date = {esc(snapshot)} LIMIT 1));"
            )
            fact_count += 1

        # Assessment values (ADI, NOAEL, ARfD, etc.) — one fact per assessment
        emitted_for_substance = set()
        for a in assessments:
            atype = a['type']
            if not atype:
                continue
            key = (atype, a.get('value'))
            if key in emitted_for_substance:
                continue
            emitted_for_substance.add(key)

            value_str = f"{a['qual']} {a['value']} {a['unit']}".strip()
            value_str = re.sub(r'\s+', ' ', value_str)
            status = f"EFSA {atype}: {value_str}"
            if a.get('population'):
                status += f" ({a['population']})"

            lines.append(
                f"INSERT OR IGNORE INTO regulatory_fact (canonical_id, jurisdiction, fact_type, status, regulation_ref, evidence_id) "
                f"VALUES ({esc(canonical_id)}, 'EFSA', 'classification', {esc(status)}, "
                f"{esc('EFSA OpenFoodTox ' + atype)}, "
                f"(SELECT id FROM fact_evidence WHERE source_url = {esc(SOURCE_URL)} AND snapshot_date = {esc(snapshot)} LIMIT 1));"
            )
            fact_count += 1

    lines.append('')
    lines.append(
        f"UPDATE ingestion_run SET finished_at = datetime('now'), status = 'completed', rows_inserted = {ingredient_count + alias_count + fact_count} "
        f"WHERE id = (SELECT id FROM ingestion_run WHERE ingester_name = {esc(INGESTER)} AND status = 'running' ORDER BY started_at DESC LIMIT 1);"
    )

    OUTPUT.write_text('\n'.join(lines))
    print('')
    print(f'Wrote {OUTPUT}')
    print(f'  ingredients: {ingredient_count}')
    print(f'  aliases:     {alias_count}')
    print(f'  facts:       {fact_count}')
    print(f'  skipped:     {skipped}')
    print(f'  file size:   {OUTPUT.stat().st_size / 1024 / 1024:.2f} MB')


if __name__ == '__main__':
    main()
